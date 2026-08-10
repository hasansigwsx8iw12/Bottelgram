from telegram import Update, ReplyKeyboardMarkup
from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    MessageHandler,
    ConversationHandler,
    ContextTypes,
    filters,
)

from datetime import datetime
from openpyxl import Workbook, load_workbook

import json
import os


# =========================================================
# إعدادات البوت
# =========================================================

TOKEN = "8403208232:AAGyopVwOIsxUJ1oisZpupHCM5HZufXJj9M"


# =========================================================
# الملفات
# =========================================================

MAINTENANCE_FILE = "maintenance_requests.json"
INSTALLATION_FILE = "installations.json"
INVENTORY_FILE = "inventory.json"
STATISTICS_FILE = "statistics.json"
EXCEL_FILE = "operations_log.xlsx"


# =========================================================
# الإحصائيات
# =========================================================

statistics = {
    "cash": 0,
    "debt": 0,
}


# =========================================================
# الموظفين
# =========================================================

EMPLOYEES = [
    "المدير ياسر",
    "نائب المدير غدير",
    "الأستاذة قمر",
    "ريم",
    "حسن",
    "دانيال",
    "محمد",
    "يزن",
    "حيدر",
]


# =========================================================
# المنتجات
# =========================================================

PRODUCTS = [
    "📡 لايت بيم M5",
    "📶 لايت AC",
    "📟 راوتر نتس",
    "📠 راوتر كودي",
]


# =========================================================
# حالات ConversationHandler
# =========================================================

(
    MAINTENANCE_NAME,
    MAINTENANCE_NATIONAL,
    MAINTENANCE_PROBLEM,
    MAINTENANCE_LOCATION,
    MAINTENANCE_PHONE,
    MAINTENANCE_REQUESTER,
    MAINTENANCE_ASSIGNEE,
    MAINTENANCE_CONFIRM,
    SEARCH_MAINTENANCE,

    ADD_PRODUCT,
    ADD_PRODUCT_QUANTITY,

    SALE_TYPE,
    PRODUCT_TYPE,
    PRODUCT_QUANTITY,
    PRODUCT_PRICE,
    PRODUCT_EMPLOYEE,
    PRODUCT_PAYMENT,

    COLLECT_AMOUNT,
    COLLECT_CUSTOMER_NAME,
    COLLECT_EMPLOYEE,
    COLLECT_CONFIRM,

    INSTALLATION_NAME,
    INSTALLATION_NATIONAL,
    INSTALLATION_TYPE,
    INSTALLATION_LOCATION,
    INSTALLATION_PHONE,
    INSTALLATION_REQUESTER,
    INSTALLATION_ASSIGNEE,
    INSTALLATION_CONFIRM,

    SEARCH_INSTALLATION,
) = range(30)


# =========================================================
# لوحات المفاتيح
# =========================================================

MAIN_KEYBOARD = [
    ["🛠 طلب صيانة"],
    ["🔍 بحث طلبات صيانة"],
    ["🌐 طلب تركيب جديد"],
    ["🔍 بحث طلبات تركيب"],
    ["📦 سحب بضاعة"],
    ["➕ إضافة بضاعة"],
    ["📊 الإحصائيات"],
    ["💰 تحصيل دين"],
]

MAIN_MARKUP = ReplyKeyboardMarkup(
    MAIN_KEYBOARD,
    resize_keyboard=True
)

BACK_MARKUP = ReplyKeyboardMarkup(
    [["🔙 رجوع"]],
    resize_keyboard=True
)

CONFIRM_MARKUP = ReplyKeyboardMarkup(
    [["✅ تأكيد", "❌ إلغاء"]],
    resize_keyboard=True
)


# =========================================================
# أدوات عامة
# =========================================================

def now():
    return datetime.now().strftime("%Y-%m-%d %H:%M")


def get_user_name(update: Update):
    user = update.effective_user

    if not user:
        return "Unknown"

    return user.full_name or user.username or str(user.id)


async def show_main_menu(update: Update):
    await update.message.reply_text(
        "🏠 القائمة الرئيسية\n\nاختر العملية التي تريدها:",
        reply_markup=MAIN_MARKUP
    )


# =========================================================
# JSON
# =========================================================

def load_json(filename, default):
    if not os.path.exists(filename):
        return default

    try:
        with open(filename, "r", encoding="utf-8") as file:
            return json.load(file)

    except Exception as e:
        print(f"JSON LOAD ERROR [{filename}]: {e}")
        return default


def save_json(filename, data):
    try:
        with open(filename, "w", encoding="utf-8") as file:
            json.dump(
                data,
                file,
                ensure_ascii=False,
                indent=4
            )

    except Exception as e:
        print(f"JSON SAVE ERROR [{filename}]: {e}")


# =========================================================
# المخزون
# =========================================================

def get_inventory():

    default_inventory = {
        "📡 لايت بيم M5": 0,
        "📶 لايت AC": 0,
        "📟 راوتر نتس": 0,
        "📠 راوتر كودي": 0,
    }

    return load_json(
        INVENTORY_FILE,
        default_inventory
    )


def update_inventory(product, quantity):

    inventory = get_inventory()

    if product not in inventory:
        inventory[product] = 0

    inventory[product] += quantity

    if inventory[product] < 0:
        inventory[product] = 0

    save_json(INVENTORY_FILE, inventory)

    return inventory[product]


def check_inventory(product, quantity):

    inventory = get_inventory()

    return inventory.get(product, 0) >= quantity


# =========================================================
# الإحصائيات
# =========================================================

def load_statistics():

    global statistics

    data = load_json(
        STATISTICS_FILE,
        {}
    )

    statistics = {
        "cash": data.get("cash", 0),
        "debt": data.get("debt", 0),
    }


def save_statistics():
    save_json(
        STATISTICS_FILE,
        statistics
    )


# =========================================================
# Excel
# =========================================================

def init_excel():

    if os.path.exists(EXCEL_FILE):
        return

    wb = Workbook()

    default_sheet = wb.active
    wb.remove(default_sheet)

    # الصيانة
    ws = wb.create_sheet("طلبات الصيانة")

    ws.append([
        "التاريخ",
        "Telegram ID",
        "اسم المستخدم",
        "اسم العميل",
        "الرقم الوطني",
        "العطل",
        "الموقع",
        "الهاتف",
        "مقدم الطلب",
        "المكلف",
        "الحالة"
    ])

    # التركيب
    ws = wb.create_sheet("طلبات التركيب")

    ws.append([
        "التاريخ",
        "Telegram ID",
        "اسم المستخدم",
        "اسم العميل",
        "الرقم الوطني",
        "نوع التركيب",
        "الموقع",
        "الهاتف",
        "مقدم الطلب",
        "المكلف بالتركيب",
        "الحالة"
    ])

    # المخزون
    ws = wb.create_sheet("المخزون")

    ws.append([
        "التاريخ",
        "Telegram ID",
        "اسم المستخدم",
        "نوع العملية",
        "المنتج",
        "الكمية",
        "المخزون بعد العملية",
        "ملاحظات"
    ])

    # البيع المباشر
    ws = wb.create_sheet("البيع المباشر")

    ws.append([
        "التاريخ",
        "Telegram ID",
        "اسم المستخدم",
        "المنتج",
        "الكمية",
        "سعر القطعة",
        "الإجمالي",
        "الموظف",
        "طريقة الدفع",
        "نوع البيع"
    ])

    # البيع دين
    ws = wb.create_sheet("البيع دين")

    ws.append([
        "التاريخ",
        "Telegram ID",
        "اسم المستخدم",
        "المنتج",
        "الكمية",
        "سعر القطعة",
        "الإجمالي",
        "الموظف",
        "الدين المتبقي",
        "ملاحظات"
    ])

    # تحصيل الديون
    ws = wb.create_sheet("تحصيل الديون")

    ws.append([
        "التاريخ",
        "Telegram ID",
        "اسم المستخدم",
        "المبلغ المحصل",
        "اسم صاحب الدين",
        "الموظف المحصل",
        "الدين المتبقي",
        "ملاحظات"
    ])

    # الإحصائيات
    ws = wb.create_sheet("الإحصائيات")

    ws.append([
        "تاريخ التحديث",
        "Telegram ID",
        "اسم المستخدم",
        "إجمالي المقبوضات",
        "إجمالي الديون",
        "مخزون لايت بيم M5",
        "مخزون لايت AC",
        "مخزون راوتر نتس",
        "مخزون راوتر كودي"
    ])

    # الموظفين
    ws = wb.create_sheet("سجل الموظفين")

    ws.append([
        "التاريخ",
        "الموظف",
        "نوع العملية",
        "المبلغ",
        "العميل",
        "ملاحظات"
    ])

    wb.save(EXCEL_FILE)

    print("✅ تم إنشاء ملف Excel")


def save_to_excel(sheet_name, row_data):

    try:

        init_excel()

        wb = load_workbook(EXCEL_FILE)

        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)

        ws = wb[sheet_name]

        ws.append(row_data)

        wb.save(EXCEL_FILE)

    except Exception as e:
        print(f"❌ Excel Error: {e}")


# =========================================================
# START
# =========================================================

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):

    context.user_data.clear()

    await update.message.reply_text(
        "👋 أهلاً وسهلاً بك في بوت KhlloNet\n\n"
        "اختر العملية من القائمة:",
        reply_markup=MAIN_MARKUP
    )


# =========================================================
# إلغاء
# =========================================================

async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):

    context.user_data.clear()

    await update.message.reply_text(
        "❌ تم إلغاء العملية",
        reply_markup=MAIN_MARKUP
    )

    return ConversationHandler.END


# =========================================================
# القائمة الرئيسية
# =========================================================

async def buttons(update: Update, context: ContextTypes.DEFAULT_TYPE):

    text = update.message.text

    # -------------------------
    # صيانة
    # -------------------------

    if text == "🛠 طلب صيانة":

        await update.message.reply_text(
            "👤 اسم العميل:",
            reply_markup=BACK_MARKUP
        )

        return MAINTENANCE_NAME

    # -------------------------
    # بحث صيانة
    # -------------------------

    if text == "🔍 بحث طلبات صيانة":

        await update.message.reply_text(
            "🔍 أدخل اسم العميل أو رقم الهاتف:",
            reply_markup=BACK_MARKUP
        )

        return SEARCH_MAINTENANCE

    # -------------------------
    # تركيب
    # -------------------------

    if text == "🌐 طلب تركيب جديد":

        await update.message.reply_text(
            "👤 اسم العميل:",
            reply_markup=BACK_MARKUP
        )

        return INSTALLATION_NAME

    # -------------------------
    # بحث تركيب
    # -------------------------

    if text == "🔍 بحث طلبات تركيب":

        await update.message.reply_text(
            "🔍 أدخل اسم العميل أو رقم الهاتف:",
            reply_markup=BACK_MARKUP
        )

        return SEARCH_INSTALLATION

    # -------------------------
    # إضافة بضاعة
    # -------------------------

    if text == "➕ إضافة بضاعة":

        keyboard = [[p] for p in PRODUCTS]
        keyboard.append(["🔙 رجوع"])

        await update.message.reply_text(
            "📦 اختر البضاعة:",
            reply_markup=ReplyKeyboardMarkup(
                keyboard,
                resize_keyboard=True
            )
        )

        return ADD_PRODUCT

    # -------------------------
    # سحب بضاعة
    # -------------------------

    if text == "📦 سحب بضاعة":

        keyboard = [
            ["🔧 بيع تركيب"],
            ["🛒 بيع مفرق"],
            ["🏪 بيع جملة"],
            ["🔙 رجوع"]
        ]

        await update.message.reply_text(
            "🛒 اختر نوع البيع:",
            reply_markup=ReplyKeyboardMarkup(
                keyboard,
                resize_keyboard=True
            )
        )

        return SALE_TYPE

    # -------------------------
    # تحصيل دين
    # -------------------------

    if text == "💰 تحصيل دين":

        await update.message.reply_text(
            "💰 أدخل المبلغ:",
            reply_markup=BACK_MARKUP
        )

        return COLLECT_AMOUNT

    # -------------------------
    # إحصائيات
    # -------------------------

    if text == "📊 الإحصائيات":

        inventory = get_inventory()

        message = f"""
📊 الإحصائيات

💰 المقبوضات: {statistics.get("cash", 0)}
💳 الديون: {statistics.get("debt", 0)}

━━━━━━━━━━━━━━

📦 المخزون

📡 لايت بيم M5:
{inventory.get("📡 لايت بيم M5", 0)}

📶 لايت AC:
{inventory.get("📶 لايت AC", 0)}

📟 راوتر نتس:
{inventory.get("📟 راوتر نتس", 0)}

📠 راوتر كودي:
{inventory.get("📠 راوتر كودي", 0)}
"""

        await update.message.reply_text(
            message,
            reply_markup=MAIN_MARKUP
        )

        return ConversationHandler.END

    # -------------------------
    # رجوع
    # -------------------------

    if text == "🔙 رجوع":

        context.user_data.clear()

        await show_main_menu(update)

        return ConversationHandler.END

    return ConversationHandler.END


# =========================================================
# الصيانة
# =========================================================

async def maintenance_name(update, context):

    if update.message.text == "🔙 رجوع":
        return await cancel(update, context)

    context.user_data["name"] = update.message.text

    await update.message.reply_text(
        "🆔 الرقم الوطني:",
        reply_markup=BACK_MARKUP
    )

    return MAINTENANCE_NATIONAL


async def maintenance_national(update, context):

    if update.message.text == "🔙 رجوع":
        return await cancel(update, context)

    context.user_data["national"] = update.message.text

    await update.message.reply_text(
        "🛠 اكتب العطل:",
        reply_markup=BACK_MARKUP
    )

    return MAINTENANCE_PROBLEM


async def maintenance_problem(update, context):

    if update.message.text == "🔙 رجوع":
        return await cancel(update, context)

    context.user_data["problem"] = update.message.text

    await update.message.reply_text(
        "📍 الموقع:",
        reply_markup=BACK_MARKUP
    )

    return MAINTENANCE_LOCATION


async def maintenance_location(update, context):

    if update.message.text == "🔙 رجوع":
        return await cancel(update, context)

    context.user_data["location"] = update.message.text

    await update.message.reply_text(
        "📞 رقم الهاتف:",
        reply_markup=BACK_MARKUP
    )

    return MAINTENANCE_PHONE


async def maintenance_phone(update, context):

    if update.message.text == "🔙 رجوع":
        return await cancel(update, context)

    context.user_data["phone"] = update.message.text

    await update.message.reply_text(
        "🙋 اسم مقدم الطلب:",
        reply_markup=BACK_MARKUP
    )

    return MAINTENANCE_REQUESTER


async def maintenance_requester(update, context):

    if update.message.text == "🔙 رجوع":
        return await cancel(update, context)

    context.user_data["requester"] = update.message.text

    keyboard = [[employee] for employee in EMPLOYEES]
    keyboard.append(["🔙 رجوع"])

    await update.message.reply_text(
        "👨‍🔧 اختر المكلف بالصيانة:",
        reply_markup=ReplyKeyboardMarkup(
            keyboard,
            resize_keyboard=True
        )
    )

    return MAINTENANCE_ASSIGNEE


async def maintenance_assignee(update, context):

    if update.message.text == "🔙 رجوع":
        return await cancel(update, context)

    context.user_data["assignee"] = update.message.text

    user_id = update.effective_user.id

    data = context.user_data

    summary = f"""
📋 ملخص طلب الصيانة

👤 الاسم:
{data["name"]}

🆔 الرقم الوطني:
{data["national"]}

🛠 العطل:
{data["problem"]}

📍 الموقع:
{data["location"]}

📞 الهاتف:
{data["phone"]}

🙋 مقدم الطلب:
{data["requester"]}

👨‍🔧 المكلف:
{data["assignee"]}

🆔 Telegram ID:
{user_id}

هل تريد تسجيل الطلب؟
"""

    await update.message.reply_text(
        summary,
        reply_markup=CONFIRM_MARKUP
    )

    return MAINTENANCE_CONFIRM


async def maintenance_confirm(update, context):

    if update.message.text == "❌ إلغاء":
        return await cancel(update, context)

    if update.message.text != "✅ تأكيد":
        return MAINTENANCE_CONFIRM

    user_id = update.effective_user.id
    data = context.user_data

    request = {
        "name": data["name"],
        "national": data["national"],
        "problem": data["problem"],
        "location": data["location"],
        "phone": data["phone"],
        "requester": data["requester"],
        "assignee": data["assignee"],
        "telegram_id": user_id,
        "date": now(),
        "status": "pending",
    }

    requests = load_json(
        MAINTENANCE_FILE,
        []
    )

    requests.append(request)

    save_json(
        MAINTENANCE_FILE,
        requests
    )

    save_to_excel(
        "طلبات الصيانة",
        [
            request["date"],
            user_id,
            get_user_name(update),
            request["name"],
            request["national"],
            request["problem"],
            request["location"],
            request["phone"],
            request["requester"],
            request["assignee"],
            request["status"],
        ]
    )

    context.user_data.clear()

    await update.message.reply_text(
        "✅ تم تسجيل طلب الصيانة بنجاح\n\n"
        f"👨‍🔧 المكلف: {request['assignee']}",
        reply_markup=MAIN_MARKUP
    )

    return ConversationHandler.END


# =========================================================
# بحث الصيانة
# =========================================================

async def search_maintenance(update, context):

    if update.message.text == "🔙 رجوع":
        return await cancel(update, context)

    term = update.message.text.strip().lower()

    requests = load_json(
        MAINTENANCE_FILE,
        []
    )

    found = []

    for request in requests:

        name = str(request.get("name", "")).lower()
        phone = str(request.get("phone", "")).lower()

        if term in name or term in phone:
            found.append(request)

    if not found:

        await update.message.reply_text(
            "❌ لم يتم العثور على أي طلب.",
            reply_markup=MAIN_MARKUP
        )

        return ConversationHandler.END

    message = ""

    for request in found:

        message += f"""
👤 {request.get("name", "-")}

🛠 العطل:
{request.get("problem", "-")}

📍 الموقع:
{request.get("location", "-")}

📞 الهاتف:
{request.get("phone", "-")}

👨‍🔧 المكلف:
{request.get("assignee", "-")}

📅 التاريخ:
{request.get("date", "-")}

📌 الحالة:
{request.get("status", "-")}

━━━━━━━━━━━━━━
"""

    await update.message.reply_text(
        message[:4000],
        reply_markup=MAIN_MARKUP
    )

    return ConversationHandler.END


# =========================================================
# التركيب
# =========================================================

async def installation_name(update, context):

    if update.message.text == "🔙 رجوع":
        return await cancel(update, context)

    context.user_data["inst_name"] = update.message.text

    await update.message.reply_text(
        "🆔 الرقم الوطني:",
        reply_markup=BACK_MARKUP
    )

    return INSTALLATION_NATIONAL


async def installation_national(update, context):

    if update.message.text == "🔙 رجوع":
        return await cancel(update, context)

    context.user_data["inst_national"] = update.message.text

    keyboard = [[p] for p in PRODUCTS]
    keyboard.append(["🔙 رجوع"])

    await update.message.reply_text(
        "🔧 نوع التركيب:",
        reply_markup=ReplyKeyboardMarkup(
            keyboard,
            resize_keyboard=True
        )
    )

    return INSTALLATION_TYPE


async def installation_type(update, context):

    if update.message.text == "🔙 رجوع":
        return await cancel(update, context)

    context.user_data["inst_type"] = update.message.text

    await update.message.reply_text(
        "📍 الموقع:",
        reply_markup=BACK_MARKUP
    )

    return INSTALLATION_LOCATION


async def installation_location(update, context):

    if update.message.text == "🔙 رجوع":
        return await cancel(update, context)

    context.user_data["inst_location"] = update.message.text

    await update.message.reply_text(
        "📞 رقم الهاتف:",
        reply_markup=BACK_MARKUP
    )

    return INSTALLATION_PHONE


async def installation_phone(update, context):

    if update.message.text == "🔙 رجوع":
        return await cancel(update, context)

    context.user_data["inst_phone"] = update.message.text

    await update.message.reply_text(
        "🙋 مقدم الطلب:",
        reply_markup=BACK_MARKUP
    )

    return INSTALLATION_REQUESTER


async def installation_requester(update, context):

    if update.message.text == "🔙 رجوع":
        return await cancel(update, context)

    context.user_data["inst_requester"] = update.message.text

    keyboard = [[employee] for employee in EMPLOYEES]
    keyboard.append(["🔙 رجوع"])

    await update.message.reply_text(
        "👨‍🔧 اختر المكلف بالتركيب:",
        reply_markup=ReplyKeyboardMarkup(
            keyboard,
            resize_keyboard=True
        )
    )

    return INSTALLATION_ASSIGNEE


async def installation_assignee(update, context):

    if update.message.text == "🔙 رجوع":
        return await cancel(update, context)

    context.user_data["inst_assignee"] = update.message.text

    data = context.user_data

    summary = f"""
📋 ملخص طلب التركيب

👤 الاسم:
{data["inst_name"]}

🆔 الرقم الوطني:
{data["inst_national"]}

🔧 النوع:
{data["inst_type"]}

📍 الموقع:
{data["inst_location"]}

📞 الهاتف:
{data["inst_phone"]}

🙋 مقدم الطلب:
{data["inst_requester"]}

👨‍🔧 المكلف:
{data["inst_assignee"]}

هل تريد تسجيل الطلب؟
"""

    await update.message.reply_text(
        summary,
        reply_markup=CONFIRM_MARKUP
    )

    return INSTALLATION_CONFIRM


async def installation_confirm(update, context):

    if update.message.text == "❌ إلغاء":
        return await cancel(update, context)

    if update.message.text != "✅ تأكيد":
        return INSTALLATION_CONFIRM

    user_id = update.effective_user.id
    data = context.user_data

    installation = {
        "name": data["inst_name"],
        "national": data["inst_national"],
        "install_type": data["inst_type"],
        "location": data["inst_location"],
        "phone": data["inst_phone"],
        "requester": data["inst_requester"],
        "assigned_to": data["inst_assignee"],
        "telegram_id": user_id,
        "date": now(),
        "status": "pending",
    }

    installations = load_json(
        INSTALLATION_FILE,
        []
    )

    installations.append(installation)

    save_json(
        INSTALLATION_FILE,
        installations
    )

    save_to_excel(
        "طلبات التركيب",
        [
            installation["date"],
            user_id,
            get_user_name(update),
            installation["name"],
            installation["national"],
            installation["install_type"],
            installation["location"],
            installation["phone"],
            installation["requester"],
            installation["assigned_to"],
            installation["status"],
        ]
    )

    context.user_data.clear()

    await update.message.reply_text(
        "✅ تم تسجيل طلب التركيب بنجاح",
        reply_markup=MAIN_MARKUP
    )

    return ConversationHandler.END


# =========================================================
# بحث التركيب
# =========================================================

async def search_installation(update, context):

    if update.message.text == "🔙 رجوع":
        return await cancel(update, context)

    term = update.message.text.strip().lower()

    installations = load_json(
        INSTALLATION_FILE,
        []
    )

    found = []

    for installation in installations:

        name = str(
            installation.get("name", "")
        ).lower()

        phone = str(
            installation.get("phone", "")
        ).lower()

        if term in name or term in phone:
            found.append(installation)

    if not found:

        await update.message.reply_text(
            "❌ لم يتم العثور على أي طلب.",
            reply_markup=MAIN_MARKUP
        )

        return ConversationHandler.END

    message = ""

    for installation in found:

        message += f"""
👤 {installation.get("name", "-")}

🔧 النوع:
{installation.get("install_type", "-")}

📍 الموقع:
{installation.get("location", "-")}

📞 الهاتف:
{installation.get("phone", "-")}

👨‍🔧 المكلف:
{installation.get("assigned_to", "-")}

📅 التاريخ:
{installation.get("date", "-")}

📌 الحالة:
{installation.get("status", "-")}

━━━━━━━━━━━━━━
"""

    await update.message.reply_text(
        message[:4000],
        reply_markup=MAIN_MARKUP
    )

    return ConversationHandler.END


# =========================================================
# إضافة بضاعة
# =========================================================

async def add_product(update, context):

    if update.message.text == "🔙 رجوع":
        return await cancel(update, context)

    context.user_data["product"] = update.message.text

    await update.message.reply_text(
        "🔢 أدخل الكمية:",
        reply_markup=BACK_MARKUP
    )

    return ADD_PRODUCT_QUANTITY


async def add_product_quantity(update, context):

    if update.message.text == "🔙 رجوع":
        return await cancel(update, context)

    try:

        quantity = int(update.message.text)

        if quantity <= 0:
            raise ValueError

        product = context.user_data["product"]

        new_quantity = update_inventory(
            product,
            quantity
        )

        save_to_excel(
            "المخزون",
            [
                now(),
                update.effective_user.id,
                get_user_name(update),
                "إضافة",
                product,
                f"+{quantity}",
                new_quantity,
                ""
            ]
        )

        context.user_data.clear()

        await update.message.reply_text(
            f"✅ تمت إضافة البضاعة\n\n"
            f"📦 المنتج: {product}\n"
            f"🔢 الكمية المضافة: {quantity}\n"
            f"📊 المخزون الحالي: {new_quantity}",
            reply_markup=MAIN_MARKUP
        )

        return ConversationHandler.END

    except ValueError:

        await update.message.reply_text(
            "❌ أدخل رقم صحيح أكبر من صفر:",
            reply_markup=BACK_MARKUP
        )

        return ADD_PRODUCT_QUANTITY


# =========================================================
# نوع البيع
# =========================================================

async def sale_type(update, context):

    if update.message.text == "🔙 رجوع":
        return await cancel(update, context)

    context.user_data["sale_type"] = update.message.text

    keyboard = [[p] for p in PRODUCTS]
    keyboard.append(["🔙 رجوع"])

    await update.message.reply_text(
        "📦 اختر المنتج:",
        reply_markup=ReplyKeyboardMarkup(
            keyboard,
            resize_keyboard=True
        )
    )

    return PRODUCT_TYPE


# =========================================================
# اختيار المنتج
# =========================================================

async def product_type(update, context):

    if update.message.text == "🔙 رجوع":
        return await cancel(update, context)

    product = update.message.text

    inventory = get_inventory()

    available = inventory.get(product, 0)

    if available <= 0:

        await update.message.reply_text(
            "❌ هذا المنتج غير متوفر في المخزون.",
            reply_markup=MAIN_MARKUP
        )

        return ConversationHandler.END

    context.user_data["product"] = product

    await update.message.reply_text(
        f"🔢 الكمية\n\n"
        f"📦 المتوفر: {available}",
        reply_markup=BACK_MARKUP
    )

    return PRODUCT_QUANTITY


# =========================================================
# كمية المنتج
# =========================================================

async def product_quantity(update, context):

    if update.message.text == "🔙 رجوع":
        return await cancel(update, context)

    try:

        quantity = int(update.message.text)

        if quantity <= 0:
            raise ValueError

        product = context.user_data["product"]

        if not check_inventory(product, quantity):

            await update.message.reply_text(
                "❌ الكمية المطلوبة غير متوفرة.",
                reply_markup=BACK_MARKUP
            )

            return PRODUCT_QUANTITY

        context.user_data["quantity"] = quantity

        await update.message.reply_text(
            "💰 سعر القطعة:",
            reply_markup=BACK_MARKUP
        )

        return PRODUCT_PRICE

    except ValueError:

        await update.message.reply_text(
            "❌ أدخل كمية صحيحة:",
            reply_markup=BACK_MARKUP
        )

        return PRODUCT_QUANTITY


# =========================================================
# سعر المنتج
# =========================================================

async def product_price(update, context):

    if update.message.text == "🔙 رجوع":
        return await cancel(update, context)

    try:

        price = float(update.message.text)

        if price < 0:
            raise ValueError

        context.user_data["price"] = price

        keyboard = [[employee] for employee in EMPLOYEES]
        keyboard.append(["🔙 رجوع"])

        await update.message.reply_text(
            "👨‍💼 اختر الموظف:",
            reply_markup=ReplyKeyboardMarkup(
                keyboard,
                resize_keyboard=True
            )
        )

        return PRODUCT_EMPLOYEE

    except ValueError:

        await update.message.reply_text(
            "❌ أدخل سعر صحيح:",
            reply_markup=BACK_MARKUP
        )

        return PRODUCT_PRICE


# =========================================================
# الموظف
# =========================================================

async def product_employee(update, context):

    if update.message.text == "🔙 رجوع":
        return await cancel(update, context)

    context.user_data["employee"] = update.message.text

    keyboard = [
        ["💰 نقداً"],
        ["📝 دين"],
        ["🔙 رجوع"]
    ]

    await update.message.reply_text(
        "💳 طريقة الدفع:",
        reply_markup=ReplyKeyboardMarkup(
            keyboard,
            resize_keyboard=True
        )
    )

    return PRODUCT_PAYMENT


# =========================================================
# الدفع
# =========================================================

async def product_payment(update, context):

    if update.message.text == "🔙 رجوع":
        return await cancel(update, context)

    payment = update.message.text

    if payment not in ["💰 نقداً", "📝 دين"]:
        return PRODUCT_PAYMENT

    product = context.user_data["product"]
    quantity = context.user_data["quantity"]
    price = context.user_data["price"]
    employee = context.user_data["employee"]
    sale = context.user_data["sale_type"]

    total = quantity * price

    new_quantity = update_inventory(
        product,
        -quantity
    )

    user_id = update.effective_user.id

    if payment == "💰 نقداً":

        statistics["cash"] += total

        sheet = "البيع المباشر"

    else:

        statistics["debt"] += total

        sheet = "البيع دين"

    save_statistics()

    save_to_excel(
        sheet,
        [
            now(),
            user_id,
            get_user_name(update),
            product,
            quantity,
            price,
            total,
            employee,
            payment,
            sale,
        ]
    )

    save_to_excel(
        "المخزون",
        [
            now(),
            user_id,
            get_user_name(update),
            "سحب",
            product,
            f"-{quantity}",
            new_quantity,
            employee,
        ]
    )

    context.user_data.clear()

    await update.message.reply_text(
        f"""
✅ تمت عملية البيع

📦 المنتج:
{product}

🔢 الكمية:
{quantity}

💰 سعر القطعة:
{price}

💵 الإجمالي:
{total}

👨‍💼 الموظف:
{employee}

💳 الدفع:
{payment}

📦 المخزون المتبقي:
{new_quantity}
""",
        reply_markup=MAIN_MARKUP
    )

    return ConversationHandler.END


# =========================================================
# تحصيل الدين
# =========================================================

async def collect_amount(update, context):

    if update.message.text == "🔙 رجوع":
        return await cancel(update, context)

    try:

        amount = float(update.message.text)

        if amount <= 0:
            raise ValueError

        context.user_data["collect_amount"] = amount

        await update.message.reply_text(
            "👤 اسم صاحب الدين:",
            reply_markup=BACK_MARKUP
        )

        return COLLECT_CUSTOMER_NAME

    except ValueError:

        await update.message.reply_text(
            "❌ أدخل مبلغ صحيح:",
            reply_markup=BACK_MARKUP
        )

        return COLLECT_AMOUNT


async def collect_customer_name(update, context):

    if update.message.text == "🔙 رجوع":
        return await cancel(update, context)

    context.user_data["collect_customer"] = update.message.text

    keyboard = [[employee] for employee in EMPLOYEES]
    keyboard.append(["🔙 رجوع"])

    await update.message.reply_text(
        "👨‍💼 اختر الموظف المحصل:",
        reply_markup=ReplyKeyboardMarkup(
            keyboard,
            resize_keyboard=True
        )
    )

    return COLLECT_EMPLOYEE


async def collect_employee(update, context):

    if update.message.text == "🔙 رجوع":
        return await cancel(update, context)

    context.user_data["collect_employee"] = update.message.text

    data = context.user_data

    summary = f"""
📋 تحصيل دين

💰 المبلغ:
{data["collect_amount"]}

👤 صاحب الدين:
{data["collect_customer"]}

👨‍💼 الموظف:
{data["collect_employee"]}

هل تريد تأكيد العملية؟
"""

    await update.message.reply_text(
        summary,
        reply_markup=CONFIRM_MARKUP
    )

    return COLLECT_CONFIRM


async def collect_confirm(update, context):

    if update.message.text == "❌ إلغاء":
        return await cancel(update, context)

    if update.message.text != "✅ تأكيد":
        return COLLECT_CONFIRM

    amount = context.user_data["collect_amount"]

    current_debt = statistics.get(
        "debt",
        0
    )

    if amount > current_debt:

        await update.message.reply_text(
            f"❌ الدين الحالي أقل من المبلغ المطلوب تحصيله.\n\n"
            f"💳 الدين الحالي: {current_debt}",
            reply_markup=MAIN_MARKUP
        )

        context.user_data.clear()

        return ConversationHandler.END

    statistics["debt"] -= amount
    statistics["cash"] += amount

    save_statistics()

    user_id = update.effective_user.id

    save_to_excel(
        "تحصيل الديون",
        [
            now(),
            user_id,
            get_user_name(update),
            amount,
            context.user_data["collect_customer"],
            context.user_data["collect_employee"],
            statistics["debt"],
            ""
        ]
    )

    customer = context.user_data["collect_customer"]

    context.user_data.clear()

    await update.message.reply_text(
        f"""
✅ تم تحصيل الدين بنجاح

👤 صاحب الدين:
{customer}

💰 المبلغ:
{amount}

💳 الدين المتبقي:
{statistics["debt"]}
""",
        reply_markup=MAIN_MARKUP
    )

    return ConversationHandler.END


# =========================================================
# معالجة الأخطاء
# =========================================================

async def error_handler(update, context):

    print("================================")
    print("❌ BOT ERROR")
    print(context.error)
    print("================================")


# =========================================================
# تشغيل البوت
# =========================================================

def main():

    print("================================")
    print("🚀 KhlloNet Bot")
    print("================================")

    load_statistics()
    init_excel()

    if TOKEN == "PUT_NEW_TOKEN_HERE":

        print("❌ ضع التوكن الجديد في TOKEN")
        return

    application = (
        ApplicationBuilder()
        .token(TOKEN)
        .build()
    )

    conversation_handler = ConversationHandler(

        entry_points=[
            MessageHandler(
                filters.TEXT & ~filters.COMMAND,
                buttons
            )
        ],

        states={

            # الصيانة
            MAINTENANCE_NAME: [
                MessageHandler(
                    filters.TEXT & ~filters.COMMAND,
                    maintenance_name
                )
            ],

            MAINTENANCE_NATIONAL: [
                MessageHandler(
                    filters.TEXT & ~filters.COMMAND,
                    maintenance_national
                )
            ],

            MAINTENANCE_PROBLEM: [
                MessageHandler(
                    filters.TEXT & ~filters.COMMAND,
                    maintenance_problem
                )
            ],

            MAINTENANCE_LOCATION: [
                MessageHandler(
                    filters.TEXT & ~filters.COMMAND,
                    maintenance_location
                )
            ],

            MAINTENANCE_PHONE: [
                MessageHandler(
                    filters.TEXT & ~filters.COMMAND,
                    maintenance_phone
                )
            ],

            MAINTENANCE_REQUESTER: [
                MessageHandler(
                    filters.TEXT & ~filters.COMMAND,
                    maintenance_requester
                )
            ],

            MAINTENANCE_ASSIGNEE: [
                MessageHandler(
                    filters.TEXT & ~filters.COMMAND,
                    maintenance_assignee
                )
            ],

            MAINTENANCE_CONFIRM: [
                MessageHandler(
                    filters.TEXT & ~filters.COMMAND,
                    maintenance_confirm
                )
            ],

            SEARCH_MAINTENANCE: [
                MessageHandler(
                    filters.TEXT & ~filters.COMMAND,
                    search_maintenance
                )
            ],

            # المخزون
            ADD_PRODUCT: [
                MessageHandler(
                    filters.TEXT & ~filters.COMMAND,
                    add_product
                )
            ],

            ADD_PRODUCT_QUANTITY: [
                MessageHandler(
                    filters.TEXT & ~filters.COMMAND,
                    add_product_quantity
                )
            ],

            # البيع
            SALE_TYPE: [
                MessageHandler(
                    filters.TEXT & ~filters.COMMAND,
                    sale_type
                )
            ],

            PRODUCT_TYPE: [
                MessageHandler(
                    filters.TEXT & ~filters.COMMAND,
                    product_type
                )
            ],

            PRODUCT_QUANTITY: [
                MessageHandler(
                    filters.TEXT & ~filters.COMMAND,
                    product_quantity
                )
            ],

            PRODUCT_PRICE: [
                MessageHandler(
                    filters.TEXT & ~filters.COMMAND,
                    product_price
                )
            ],

            PRODUCT_EMPLOYEE: [
                MessageHandler(
                    filters.TEXT & ~filters.COMMAND,
                    product_employee
                )
            ],

            PRODUCT_PAYMENT: [
                MessageHandler(
                    filters.TEXT & ~filters.COMMAND,
                    product_payment
                )
            ],

            # الديون
            COLLECT_AMOUNT: [
                MessageHandler(
                    filters.TEXT & ~filters.COMMAND,
                    collect_amount
                )
            ],

            COLLECT_CUSTOMER_NAME: [
                MessageHandler(
                    filters.TEXT & ~filters.COMMAND,
                    collect_customer_name
                )
            ],

            COLLECT_EMPLOYEE: [
                MessageHandler(
                    filters.TEXT & ~filters.COMMAND,
                    collect_employee
                )
            ],

            COLLECT_CONFIRM: [
                MessageHandler(
                    filters.TEXT & ~filters.COMMAND,
                    collect_confirm
                )
            ],

            # التركيب
            INSTALLATION_NAME: [
                MessageHandler(
                    filters.TEXT & ~filters.COMMAND,
                    installation_name
                )
            ],

            INSTALLATION_NATIONAL: [
                MessageHandler(
                    filters.TEXT & ~filters.COMMAND,
                    installation_national
                )
            ],

            INSTALLATION_TYPE: [
                MessageHandler(
                    filters.TEXT & ~filters.COMMAND,
                    installation_type
                )
            ],

            INSTALLATION_LOCATION: [
                MessageHandler(
                    filters.TEXT & ~filters.COMMAND,
                    installation_location
                )
            ],

            INSTALLATION_PHONE: [
                MessageHandler(
                    filters.TEXT & ~filters.COMMAND,
                    installation_phone
                )
            ],

            INSTALLATION_REQUESTER: [
                MessageHandler(
                    filters.TEXT & ~filters.COMMAND,
                    installation_requester
                )
            ],

            INSTALLATION_ASSIGNEE: [
                MessageHandler(
                    filters.TEXT & ~filters.COMMAND,
                    installation_assignee
                )
            ],

            INSTALLATION_CONFIRM: [
                MessageHandler(
                    filters.TEXT & ~filters.COMMAND,
                    installation_confirm
                )
            ],

            SEARCH_INSTALLATION: [
                MessageHandler(
                    filters.TEXT & ~filters.COMMAND,
                    search_installation
                )
            ],
        },

        fallbacks=[
            CommandHandler(
                "cancel",
                cancel
            )
        ],

        allow_reentry=True,
    )

    # /start
    application.add_handler(
        CommandHandler(
            "start",
            start
        )
    )

    # المحادثات
    application.add_handler(
        conversation_handler
    )

    # الأخطاء
    application.add_error_handler(
        error_handler
    )

    print("✅ BOT RUNNING...")
    print("📡 Waiting for Telegram messages...")

    application.run_polling(
        drop_pending_updates=True
    )


# =========================================================
# تشغيل
# =========================================================

if __name__ == "__main__":
    main()
